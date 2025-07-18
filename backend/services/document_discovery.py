"""
Document Discovery Service - Migrated from frontend web-search.ts
"""

import asyncio
import logging
import re
from typing import Dict, List, Any, AsyncGenerator, Union
from datetime import datetime
import httpx
from bs4 import BeautifulSoup
import json
import hashlib
import os
import base64
from pathlib import Path
import mimetypes
from PIL import Image
import PyPDF2
import openpyxl
import csv
import chardet

logger = logging.getLogger(__name__)

class DocumentDiscoveryService:
    def __init__(self):
        self.discovered_documents = []
        self.document_hashes = set()
        self.downloaded_documents = []
        self.trusted_sources = [
            "cisco.com", "ciscopress.com", "ine.com", "cbtnuggets.com",
            "udemy.com", "pluralsight.com", "google.com", "google.co.za", "youtube.com"
        ]
        self.search_engines = {
            "duckduckgo": "https://api.duckduckgo.com/?q={query}&format=json&no_redirect=1&no_html=1&skip_disambig=1",
            "bing": "https://www.bing.com/search?q={query}&format=rss",
        }
        self.cors_proxy = "https://api.allorigins.win/raw?url="
        
        # Create data directory if it doesn't exist
        os.makedirs("./data/documents", exist_ok=True)
        os.makedirs("./data/uploads", exist_ok=True)
        
        # Supported file types for local upload
        self.supported_file_types = {
            '.pdf': 'application/pdf',
            '.csv': 'text/csv',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.xls': 'application/vnd.ms-excel',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.txt': 'text/plain',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif'
        }
    
    async def discover_documents(self, topic: str, certification_level: str = "all", 
                               max_documents: int = 4, sources: List[str] = None,
                               use_ai_agent: bool = False) -> AsyncGenerator[Dict[str, Any], None]:
        """Main document discovery process"""
        
        if sources is None:
            sources = self.trusted_sources
        
        # Step 1: Initialize
        yield {
            "step": "initialize",
            "progress": 0,
            "message": "Initializing document discovery...",
            "status": "running"
        }
        
        await asyncio.sleep(1)
        
        # Step 2: Generate search queries
        yield {
            "step": "generate_queries",
            "progress": 20,
            "message": "Generating search queries...",
            "status": "running"
        }
        
        search_queries = await self._generate_search_queries(topic, certification_level)
        
        # Step 3: Search trusted sources
        yield {
            "step": "search_sources",
            "progress": 40,
            "message": "Searching trusted sources...",
            "status": "running"
        }
        
        all_results = []
        for i, query in enumerate(search_queries):
            results = await self._search_web(query, sources)
            all_results.extend(results)
            
            progress = 40 + (i + 1) / len(search_queries) * 30
            yield {
                "step": "search_sources",
                "progress": progress,
                "message": f"Searched {i + 1}/{len(search_queries)} queries - Found {len(all_results)} documents",
                "status": "running"
            }
        
        # Step 4: Validate and filter documents
        yield {
            "step": "validate",
            "progress": 70,
            "message": "Validating documents...",
            "status": "running"
        }
        
        validated_docs = await self._validate_documents(all_results)
        
        # Step 5: Download and organize
        yield {
            "step": "download",
            "progress": 90,
            "message": "Organizing documents...",
            "status": "running"
        }
        
        final_docs = await self._organize_documents(validated_docs, max_documents)
        self.discovered_documents = final_docs
        
        # Final result
        yield {
            "step": "completed",
            "progress": 100,
            "message": f"Discovery completed - {len(final_docs)} documents found",
            "status": "completed",
            "documents": final_docs,
            "count": len(final_docs)
        }
    
    async def _generate_search_queries(self, topic: str, cert_level: str) -> List[str]:
        """Generate diverse search queries for the topic"""
        base_queries = [
            f"{topic} cisco configuration guide",
            f"{topic} cisco troubleshooting",
            f"{topic} cisco best practices",
            f"{topic} cisco implementation examples",
            f"{topic} cisco security configuration"
        ]
        
        if cert_level != "all":
            cert_queries = [
                f"{topic} {cert_level} study guide",
                f"{topic} {cert_level} configuration",
                f"{topic} {cert_level} troubleshooting"
            ]
            base_queries.extend(cert_queries)
        
        return base_queries[:8]  # Limit to 8 queries
    
    async def _search_web(self, query: str, sources: List[str]) -> List[Dict[str, Any]]:
        """Search web sources for documents"""
        results = []
        
        try:
            # Try DuckDuckGo first
            duck_results = await self._search_duckduckgo(query)
            results.extend(duck_results)
            
            # Search specific sites
            for source in sources:
                try:
                    site_results = await self._search_site(query, source)
                    results.extend(site_results)
                except Exception as e:
                    logger.error(f"Error searching {source}: {str(e)}")
                    
        except Exception as e:
            logger.error(f"Web search error: {str(e)}")
        
        return results
    
    async def _search_duckduckgo(self, query: str) -> List[Dict[str, Any]]:
        """Search DuckDuckGo for documents"""
        try:
            url = self.search_engines["duckduckgo"].replace("{query}", query)
            async with httpx.AsyncClient() as client:
                response = await client.get(url)
                if response.status_code != 200:
                    return []
                    
                data = response.json()
                results = []
                
                # Process instant answer
                if data.get("Answer"):
                    results.append({
                        "title": data.get("Heading", "DuckDuckGo Answer"),
                        "url": data.get("AbstractURL", "#"),
                        "snippet": data.get("Answer"),
                        "source": "duckduckgo.com"
                    })
                
                # Process related topics
                if data.get("RelatedTopics"):
                    for topic in data["RelatedTopics"]:
                        if topic.get("FirstURL") and topic.get("Text"):
                            title = topic["Text"].split(" - ")[0] if " - " in topic["Text"] else topic["Text"][:100]
                            source = topic["FirstURL"].split("/")[2] if "/" in topic["FirstURL"] else "unknown"
                            results.append({
                                "title": title,
                                "url": topic["FirstURL"],
                                "snippet": topic["Text"],
                                "source": source
                            })
                
                return results
        except Exception as e:
            logger.error(f"DuckDuckGo search error: {str(e)}")
            return []
    
    async def _search_site(self, query: str, site: str) -> List[Dict[str, Any]]:
        """Search a specific site for documents"""
        # In production, this would use site-specific search APIs or web scraping
        # For now, generate mock results based on query and site
        mock_docs = await self._generate_mock_documents(query, site)
        return mock_docs
    
    async def _generate_mock_documents(self, query: str, site: str) -> List[Dict[str, Any]]:
        """Generate realistic mock documents for testing"""
        topics = query.lower()
        docs = []
        
        # Document library with comprehensive coverage
        document_library = {
            "bgp": [
                {
                    "title": f"BGP Configuration Guide - {site}",
                    "url": f"https://{site}/bgp-configuration-guide.pdf",
                    "summary": "Comprehensive guide for BGP configuration and troubleshooting on Cisco devices.",
                    "type": "PDF",
                    "size": "2.4 MB"
                },
                {
                    "title": f"BGP Security Best Practices - {site}",
                    "url": f"https://{site}/bgp-security-best-practices.pdf",
                    "summary": "Advanced BGP security configurations and threat mitigation strategies.",
                    "type": "PDF",
                    "size": "1.8 MB"
                },
                {
                    "title": f"BGP Route Reflector Design - {site}",
                    "url": f"https://{site}/bgp-route-reflector-design.pdf",
                    "summary": "Scalable BGP route reflector architectures for large networks.",
                    "type": "PDF",
                    "size": "3.2 MB"
                }
            ],
            "ospf": [
                {
                    "title": f"OSPF Implementation Guide - {site}",
                    "url": f"https://{site}/ospf-implementation-guide.pdf",
                    "summary": "Detailed OSPF implementation and design considerations for enterprise networks.",
                    "type": "PDF",
                    "size": "3.1 MB"
                },
                {
                    "title": f"OSPF Troubleshooting Handbook - {site}",
                    "url": f"https://{site}/ospf-troubleshooting.pdf",
                    "summary": "Common OSPF issues and systematic troubleshooting approaches.",
                    "type": "PDF",
                    "size": "2.7 MB"
                },
                {
                    "title": f"OSPF Area Design Principles - {site}",
                    "url": f"https://{site}/ospf-area-design.pdf",
                    "summary": "OSPF area design strategies and hierarchical network planning.",
                    "type": "PDF",
                    "size": "1.9 MB"
                }
            ],
            "mpls": [
                {
                    "title": f"MPLS VPN Configuration - {site}",
                    "url": f"https://{site}/mpls-vpn-configuration.pdf",
                    "summary": "Advanced MPLS VPN configuration and troubleshooting techniques.",
                    "type": "PDF",
                    "size": "4.2 MB"
                },
                {
                    "title": f"MPLS Traffic Engineering - {site}",
                    "url": f"https://{site}/mpls-traffic-engineering.pdf",
                    "summary": "MPLS-TE implementation for optimized network traffic flow.",
                    "type": "PDF",
                    "size": "3.5 MB"
                }
            ],
            "eigrp": [
                {
                    "title": f"EIGRP Configuration and Tuning - {site}",
                    "url": f"https://{site}/eigrp-configuration.pdf",
                    "summary": "EIGRP protocol configuration, optimization, and troubleshooting.",
                    "type": "PDF",
                    "size": "2.8 MB"
                }
            ],
            "qos": [
                {
                    "title": f"QoS Implementation Guide - {site}",
                    "url": f"https://{site}/qos-implementation.pdf",
                    "summary": "Quality of Service configuration for voice, video, and data traffic.",
                    "type": "PDF",
                    "size": "3.3 MB"
                },
                {
                    "title": f"Advanced QoS Techniques - {site}",
                    "url": f"https://{site}/advanced-qos-techniques.pdf",
                    "summary": "Advanced QoS mechanisms including CBWFQ, LLQ, and traffic shaping.",
                    "type": "PDF",
                    "size": "2.9 MB"
                }
            ],
            "security": [
                {
                    "title": f"Cisco ASA Firewall Configuration - {site}",
                    "url": f"https://{site}/asa-firewall-config.pdf",
                    "summary": "Comprehensive ASA firewall configuration and security policies.",
                    "type": "PDF",
                    "size": "4.5 MB"
                },
                {
                    "title": f"Network Security Fundamentals - {site}",
                    "url": f"https://{site}/network-security-fundamentals.pdf",
                    "summary": "Core network security concepts and Cisco security solutions.",
                    "type": "PDF",
                    "size": "3.7 MB"
                }
            ],
            "switching": [
                {
                    "title": f"VLAN and Trunking Configuration - {site}",
                    "url": f"https://{site}/vlan-trunking-config.pdf",
                    "summary": "VLAN design, trunking protocols, and inter-VLAN routing.",
                    "type": "PDF",
                    "size": "2.6 MB"
                },
                {
                    "title": f"Spanning Tree Protocol Guide - {site}",
                    "url": f"https://{site}/spanning-tree-guide.pdf",
                    "summary": "STP, RSTP, and MST configuration for loop-free switching.",
                    "type": "PDF",
                    "size": "3.0 MB"
                }
            ],
            "wireless": [
                {
                    "title": f"Cisco Wireless LAN Configuration - {site}",
                    "url": f"https://{site}/wireless-lan-config.pdf",
                    "summary": "Wireless controller and access point configuration guide.",
                    "type": "PDF",
                    "size": "3.8 MB"
                }
            ]
        }
        
        # Match topics and add relevant documents
        for topic, topic_docs in document_library.items():
            if topic in topics:
                for doc in topic_docs:
                    doc_id = hashlib.md5(f"{doc['url']}_{doc['title']}".encode()).hexdigest()[:8]
                    docs.append({
                        "id": doc_id,
                        "title": doc["title"],
                        "url": doc["url"],
                        "source": site,
                        "type": doc["type"],
                        "size": doc["size"],
                        "summary": doc["summary"],
                        "relevance": self._calculate_relevance({"title": doc["title"], "snippet": doc["summary"], "source": site}, query),
                        "downloadStatus": "pending"
                    })
        
        # Add generic result if no specific matches
        if not docs:
            doc_id = hashlib.md5(f"{query}_{site}".encode()).hexdigest()[:8]
            docs.append({
                "id": doc_id,
                "title": f"{query.title()} Configuration Guide - {site}",
                "url": f"https://{site}/{query.replace(' ', '-').lower()}-guide.pdf",
                "source": site,
                "type": "PDF",
                "size": f"{2.0 + (hash(query) % 30) / 10:.1f} MB",
                "summary": f"Comprehensive documentation and configuration examples for {query}.",
                "relevance": 0.75 + (hash(query) % 20) / 100,
                "downloadStatus": "pending"
            })
            
        # Add some latent documents randomly
        latent_docs = [
            {
                "title": f"Network Troubleshooting Methodology - {site}",
                "url": f"https://{site}/network-troubleshooting-methodology.pdf",
                "summary": "Systematic approach to network problem diagnosis and resolution.",
                "source": site,
                "type": "PDF",
                "size": "2.3 MB"
            },
            {
                "title": f"Cisco IOS Command Reference - {site}",
                "url": f"https://{site}/ios-command-reference.pdf",
                "summary": "Complete reference for Cisco IOS commands and syntax.",
                "source": site,
                "type": "PDF",
                "size": "5.1 MB"
            },
            {
                "title": f"Network Design Best Practices - {site}",
                "url": f"https://{site}/network-design-best-practices.pdf",
                "summary": "Industry best practices for scalable network architecture design.",
                "source": site,
                "type": "PDF",
                "size": "3.4 MB"
            }
        ]
        
        if len(docs) < 5 and hash(query) % 2 == 0:
            latent_doc = latent_docs[hash(query) % len(latent_docs)]
            doc_id = hashlib.md5(f"{latent_doc['url']}_{latent_doc['title']}".encode()).hexdigest()[:8]
            docs.append({
                "id": doc_id,
                "title": latent_doc["title"],
                "url": latent_doc["url"],
                "source": latent_doc["source"],
                "type": latent_doc["type"],
                "size": latent_doc["size"],
                "summary": latent_doc["summary"],
                "relevance": self._calculate_relevance({"title": latent_doc["title"], "snippet": latent_doc["summary"], "source": site}, query),
                "downloadStatus": "pending"
            })
        
        return docs
    
    def _calculate_relevance(self, result: Dict[str, str], query: str) -> float:
        """Calculate relevance score for a search result"""
        query_words = query.lower().split()
        title_words = result["title"].lower().split()
        snippet_words = result["snippet"].lower().split()
        
        score = 0
        total_words = len(query_words) if query_words else 1
        
        for word in query_words:
            if any(word in tw for tw in title_words):
                score += 0.4
            if any(word in sw for sw in snippet_words):
                score += 0.2
            if "cisco" in result["source"]:
                score += 0.1
        
        # Add some randomness for variety
        return min(score / total_words + (hash(result["title"]) % 20) / 100, 1.0)
    
    def _detect_document_type(self, url: str, title: str) -> str:
        """Detect document type based on URL and title"""
        if ".pdf" in url.lower() or "pdf" in title.lower():
            return "PDF"
        if ".doc" in url.lower() or "doc" in title.lower():
            return "DOC"
        if ".ppt" in url.lower() or "presentation" in title.lower():
            return "PPT"
        return "HTML"
    
    def _estimate_size(self, title: str) -> str:
        """Estimate document size based on title length"""
        base_size = 0.5 + (len(title) % 30) / 10  # 0.5 - 3.5 MB
        return f"{base_size:.1f} MB"
    
    async def _validate_documents(self, documents: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Validate and deduplicate documents"""
        validated = []
        seen_urls = set()
        
        for doc in documents:
            # Check for duplicates
            if doc["url"] in seen_urls:
                continue
            
            seen_urls.add(doc["url"])
            
            # In production, validate URL accessibility
            # For now, just mark as validated
            doc["validated"] = True
            doc["validation_date"] = datetime.now().isoformat()
            
            # Generate document hash for tracking
            doc_hash = self._generate_document_hash(doc)
            doc["document_hash"] = doc_hash
            
            # Check if already in document_hashes
            if doc_hash not in self.document_hashes:
                self.document_hashes.add(doc_hash)
                validated.append(doc)
        
        return validated
    
    def _generate_document_hash(self, document: Dict[str, Any]) -> str:
        """Generate a unique hash for document to prevent duplicates"""
        hash_string = f"{document['url']}_{document['title']}_{document['source']}"
        return hashlib.md5(hash_string.encode()).hexdigest()[:16]
    
    async def _organize_documents(self, documents: List[Dict[str, Any]], max_docs: int) -> List[Dict[str, Any]]:
        """Organize and limit documents"""
        # Sort by relevance
        sorted_docs = sorted(documents, key=lambda x: x["relevance"], reverse=True)
        
        # Limit to max_documents
        final_docs = sorted_docs[:max_docs]
        
        # Add metadata
        for i, doc in enumerate(final_docs):
            doc["rank"] = i + 1
            doc["discovered_at"] = datetime.now().isoformat()
            doc["page_references"] = [i * 10 + j for j in range(1, 4)]  # Mock page references
        
        return final_docs
    
    async def get_documents(self) -> List[Dict[str, Any]]:
        """Get list of discovered documents"""
        return self.discovered_documents
    
    async def download_document(self, document_id: str, document: Dict[str, Any] = None) -> Dict[str, Any]:
        """Download a specific document"""
        # Find document if not provided
        if not document:
            doc = next((d for d in self.discovered_documents if d["id"] == document_id), None)
            if not doc:
                raise ValueError(f"Document {document_id} not found")
        else:
            doc = document
        
        # Check if document already exists
        if self._is_document_already_downloaded(doc):
            logger.info(f"Document already exists: {doc['title']}")
            doc["downloadStatus"] = "completed"
            doc["downloadProgress"] = 100
            doc["downloaded_at"] = datetime.now().isoformat()
            return doc
        
        # Simulate download process
        doc["downloadStatus"] = "downloading"
        doc["downloadProgress"] = 0
        
        try:
            # In production, actually download the document
            # For now, simulate the download process
            for progress in range(0, 101, 10):
                await asyncio.sleep(0.2)
                doc["downloadProgress"] = progress
            
            # Save document info
            doc["downloadStatus"] = "completed"
            doc["downloaded_at"] = datetime.now().isoformat()
            doc["local_path"] = f"./data/documents/{document_id}.pdf"  # Mock local path
            
            # Add to downloaded documents list
            self.downloaded_documents.append(doc)
            
            return doc
            
        except Exception as e:
            logger.error(f"Download failed for {doc['title']}: {str(e)}")
            doc["downloadStatus"] = "failed"
            doc["download_error"] = str(e)
            return doc
    
    def _is_document_already_downloaded(self, document: Dict[str, Any]) -> bool:
        """Check if document is already downloaded"""
        doc_hash = document.get("document_hash") or self._generate_document_hash(document)
        return any(d.get("document_hash") == doc_hash for d in self.downloaded_documents)
    
    def get_downloaded_documents(self) -> List[Dict[str, Any]]:
        """Get list of downloaded documents"""
        return self.downloaded_documents
    
    def get_downloaded_document(self, document_id: str) -> Dict[str, Any]:
        """Get a specific downloaded document"""
        doc = next((d for d in self.downloaded_documents if d["id"] == document_id), None)
        if not doc:
            raise ValueError(f"Downloaded document {document_id} not found")
        return doc
    
    def clear_downloaded_documents(self) -> None:
        """Clear all downloaded documents"""
        self.downloaded_documents = []
        logger.info("Cleared all downloaded documents")
    
    def get_unprocessed_documents(self) -> List[Dict[str, Any]]:
        """Get list of unprocessed documents"""
        return [doc for doc in self.downloaded_documents if not doc.get("is_processed", False)]
    
    def mark_document_as_processed(self, document_id: str) -> None:
        """Mark a document as processed"""
        for doc in self.downloaded_documents:
            if doc["id"] == document_id:
                doc["is_processed"] = True
                doc["processed_at"] = datetime.now().isoformat()
                logger.info(f"Marked document {document_id} as processed")
                break
    
    async def process_local_files(self, files_data: List[Dict[str, Any]], 
                                 directory_path: str = None) -> AsyncGenerator[Dict[str, Any], None]:
        """Process local files and/or directory"""
        
        yield {
            "step": "initialize",
            "progress": 0,
            "message": "Initializing local file processing...",
            "status": "running"
        }
        
        await asyncio.sleep(0.5)
        
        processed_files = []
        total_files = len(files_data) + (1 if directory_path else 0)
        current_file = 0
        
        # Process uploaded files
        for file_data in files_data:
            current_file += 1
            progress = (current_file / total_files) * 80  # Reserve 20% for final steps
            
            yield {
                "step": "process_file",
                "progress": progress,
                "message": f"Processing file {current_file}/{len(files_data)}: {file_data.get('name', 'Unknown')}",
                "status": "running",
                "current_file": file_data.get('name')
            }
            
            try:
                processed_file = await self._process_single_file(file_data)
                if processed_file:
                    processed_files.append(processed_file)
                    
            except Exception as e:
                logger.error(f"Error processing file {file_data.get('name')}: {str(e)}")
                yield {
                    "step": "process_file",
                    "progress": progress,
                    "message": f"Error processing {file_data.get('name')}: {str(e)}",
                    "status": "error",
                    "current_file": file_data.get('name')
                }
        
        # Process directory if specified
        if directory_path:
            current_file += 1
            progress = (current_file / total_files) * 80
            
            yield {
                "step": "process_directory",
                "progress": progress,
                "message": f"Scanning directory: {directory_path}",
                "status": "running"
            }
            
            try:
                directory_files = await self._process_directory(directory_path)
                processed_files.extend(directory_files)
                
            except Exception as e:
                logger.error(f"Error processing directory {directory_path}: {str(e)}")
                yield {
                    "step": "process_directory",
                    "progress": progress,
                    "message": f"Error processing directory: {str(e)}",
                    "status": "error"
                }
        
        # Finalize processing
        yield {
            "step": "finalize",
            "progress": 90,
            "message": "Finalizing processed files...",
            "status": "running"
        }
        
        await asyncio.sleep(0.5)
        
        # Add to discovered documents
        self.discovered_documents.extend(processed_files)
        
        # Final result
        yield {
            "step": "completed",
            "progress": 100,
            "message": f"Local file processing completed - {len(processed_files)} files processed",
            "status": "completed",
            "documents": processed_files,
            "count": len(processed_files)
        }
    
    async def _process_single_file(self, file_data: Dict[str, Any]) -> Dict[str, Any]:
        """Process a single uploaded file"""
        
        file_name = file_data.get('name', 'unknown')
        file_content = file_data.get('content')  # Base64 encoded content
        file_type = file_data.get('type', '')
        file_size = file_data.get('size', 0)
        
        # Generate unique ID
        file_id = hashlib.md5(f"{file_name}_{datetime.now().isoformat()}".encode()).hexdigest()[:12]
        
        # Determine file extension and type
        file_path = Path(file_name)
        file_extension = file_path.suffix.lower()
        
        if file_extension not in self.supported_file_types:
            logger.warning(f"Unsupported file type: {file_extension}")
            return None
        
        # Save file locally
        local_path = f"./data/uploads/{file_id}_{file_name}"
        
        try:
            if file_content:
                # Decode base64 content and save
                file_bytes = base64.b64decode(file_content)
                with open(local_path, 'wb') as f:
                    f.write(file_bytes)
            
            # Extract text content based on file type
            extracted_text = await self._extract_text_from_file(local_path, file_extension)
            
            # Create document entry
            document = {
                "id": f"local_{file_id}",
                "title": file_name,
                "source": "local-upload",
                "type": self._get_document_type(file_extension),
                "size": self._format_file_size(file_size),
                "relevance": 1.0,  # Local files are always 100% relevant
                "downloadStatus": "completed",
                "downloadProgress": 100,
                "url": local_path,
                "summary": f"Local file: {file_name}. {extracted_text[:200]}..." if extracted_text else f"Local file: {file_name}",
                "local_path": local_path,
                "file_extension": file_extension,
                "mime_type": self.supported_file_types.get(file_extension, 'application/octet-stream'),
                "extracted_text": extracted_text,
                "is_local_file": True,
                "uploaded_at": datetime.now().isoformat()
            }
            
            return document
            
        except Exception as e:
            logger.error(f"Error processing file {file_name}: {str(e)}")
            # Clean up file if it was created
            if os.path.exists(local_path):
                os.remove(local_path)
            return None
    
    async def _process_directory(self, directory_path: str) -> List[Dict[str, Any]]:
        """Process all supported files in a directory"""
        
        if not os.path.exists(directory_path):
            raise ValueError(f"Directory does not exist: {directory_path}")
        
        processed_files = []
        directory = Path(directory_path)
        
        # Find all supported files in directory
        for file_path in directory.rglob('*'):
            if file_path.is_file() and file_path.suffix.lower() in self.supported_file_types:
                try:
                    # Read file content
                    with open(file_path, 'rb') as f:
                        file_content = base64.b64encode(f.read()).decode('utf-8')
                    
                    file_data = {
                        'name': file_path.name,
                        'content': file_content,
                        'type': self.supported_file_types.get(file_path.suffix.lower(), ''),
                        'size': file_path.stat().st_size
                    }
                    
                    processed_file = await self._process_single_file(file_data)
                    if processed_file:
                        processed_files.append(processed_file)
                        
                except Exception as e:
                    logger.error(f"Error processing file {file_path}: {str(e)}")
                    continue
        
        return processed_files
    
    async def _extract_text_from_file(self, file_path: str, file_extension: str) -> str:
        """Extract text content from various file types"""
        
        try:
            if file_extension == '.pdf':
                return await self._extract_text_from_pdf(file_path)
            elif file_extension == '.csv':
                return await self._extract_text_from_csv(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                return await self._extract_text_from_excel(file_path)
            elif file_extension == '.txt':
                return await self._extract_text_from_txt(file_path)
            elif file_extension in ['.jpg', '.jpeg', '.png', '.gif']:
                return await self._extract_text_from_image(file_path)
            elif file_extension in ['.doc', '.docx']:
                return await self._extract_text_from_word(file_path)
            else:
                return "Text extraction not supported for this file type."
                
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {str(e)}")
            return f"Error extracting text: {str(e)}"
    
    async def _extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        text_content = []
        
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text_content.append(page.extract_text())
            
            return '\n'.join(text_content)
            
        except Exception as e:
            logger.error(f"Error reading PDF {file_path}: {str(e)}")
            return f"PDF processing error: {str(e)}"
    
    async def _extract_text_from_csv(self, file_path: str) -> str:
        """Extract text from CSV file"""
        text_content = []
        
        try:
            # Detect encoding
            with open(file_path, 'rb') as file:
                raw_data = file.read()
                encoding = chardet.detect(raw_data)['encoding'] or 'utf-8'
            
            with open(file_path, 'r', encoding=encoding) as file:
                csv_reader = csv.reader(file)
                for row_num, row in enumerate(csv_reader):
                    if row_num == 0:
                        text_content.append(f"Headers: {', '.join(row)}")
                    else:
                        text_content.append(f"Row {row_num}: {', '.join(row)}")
                    
                    # Limit to first 100 rows for performance
                    if row_num >= 100:
                        text_content.append("... (truncated for performance)")
                        break
            
            return '\n'.join(text_content)
            
        except Exception as e:
            logger.error(f"Error reading CSV {file_path}: {str(e)}")
            return f"CSV processing error: {str(e)}"
    
    async def _extract_text_from_excel(self, file_path: str) -> str:
        """Extract text from Excel file"""
        text_content = []
        
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"Sheet: {sheet_name}")
                
                for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if row_num == 1:
                        text_content.append(f"Headers: {', '.join(str(cell) for cell in row if cell is not None)}")
                    else:
                        row_text = ', '.join(str(cell) for cell in row if cell is not None)
                        if row_text.strip():
                            text_content.append(f"Row {row_num}: {row_text}")
                    
                    # Limit to first 100 rows per sheet
                    if row_num >= 100:
                        text_content.append("... (truncated for performance)")
                        break
                
                text_content.append("")  # Empty line between sheets
            
            workbook.close()
            return '\n'.join(text_content)
            
        except Exception as e:
            logger.error(f"Error reading Excel {file_path}: {str(e)}")
            return f"Excel processing error: {str(e)}"
    
    async def _extract_text_from_txt(self, file_path: str) -> str:
        """Extract text from text file"""
        try:
            # Detect encoding
            with open(file_path, 'rb') as file:
                raw_data = file.read()
                encoding = chardet.detect(raw_data)['encoding'] or 'utf-8'
            
            with open(file_path, 'r', encoding=encoding) as file:
                return file.read()
                
        except Exception as e:
            logger.error(f"Error reading text file {file_path}: {str(e)}")
            return f"Text file processing error: {str(e)}"
    
    async def _extract_text_from_image(self, file_path: str) -> str:
        """Extract text from image file (basic metadata)"""
        try:
            with Image.open(file_path) as img:
                # For now, just return image metadata
                # In production, you might want to use OCR (like Tesseract)
                return f"Image file: {img.format}, Size: {img.size}, Mode: {img.mode}"
                
        except Exception as e:
            logger.error(f"Error reading image {file_path}: {str(e)}")
            return f"Image processing error: {str(e)}"
    
    async def _extract_text_from_word(self, file_path: str) -> str:
        """Extract text from Word document"""
        # Note: This would require python-docx for .docx files
        # For now, return a placeholder
        return f"Word document processing not fully implemented. File: {os.path.basename(file_path)}"
    
    def _get_document_type(self, file_extension: str) -> str:
        """Get document type based on file extension"""
        type_mapping = {
            '.pdf': 'PDF',
            '.csv': 'CSV',
            '.xlsx': 'Excel',
            '.xls': 'Excel',
            '.doc': 'Word',
            '.docx': 'Word',
            '.txt': 'Text',
            '.jpg': 'Image',
            '.jpeg': 'Image',
            '.png': 'Image',
            '.gif': 'Image'
        }
        return type_mapping.get(file_extension, 'Document')
    
    def _format_file_size(self, size_bytes: int) -> str:
        """Format file size in human readable format"""
        if size_bytes == 0:
            return "0 B"
        
        size_names = ["B", "KB", "MB", "GB"]
        i = 0
        size = float(size_bytes)
        
        while size >= 1024.0 and i < len(size_names) - 1:
            size /= 1024.0
            i += 1
        
        return f"{size:.1f} {size_names[i]}"
    
    async def get_local_files(self) -> List[Dict[str, Any]]:
        """Get list of processed local files"""
        return [doc for doc in self.discovered_documents if doc.get("is_local_file", False)]